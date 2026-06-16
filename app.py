import os
import json
import re
import platform
import pytesseract
import webbrowser
import threading
from datetime import datetime
from PIL import Image, ImageFilter, ImageEnhance
from flask import Flask, request, jsonify, send_file, session, render_template
from werkzeug.security import generate_password_hash, check_password_hash
import spacy
from langdetect import detect, LangDetectException
from deep_translator import GoogleTranslator
from pdf2image import convert_from_path
import PyPDF2
import docx
from PIL import Image
import numpy as np
import io

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    REPORTLAB_OK = True
except ImportError:
    REPORTLAB_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False

nlp_en = spacy.load("en_core_web_sm")
app = Flask(__name__)
app.secret_key = "mcqportal_secret_key_2025"

if platform.system() == "Windows":
    pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
elif platform.system() == "Darwin":
    pytesseract.pytesseract.tesseract_cmd = "/opt/homebrew/bin/tesseract"

print("Tesseract:", pytesseract.pytesseract.tesseract_cmd)

try:
    print("Available Languages:", pytesseract.get_languages())
except Exception as e:
    print("Language Load Error:", e)


def find_separator(image_path):
    img = Image.open(image_path).convert("L")
    arr = np.array(img)
    score = np.sum(255 - arr, axis=0)
    width = arr.shape[1]
    start = int(width * 0.35)
    end = int(width * 0.65)
    separator = start + np.argmax(score[start:end])
    return separator


DATA_FILE = "course_data.json"


def load_data():
    if not os.path.exists(DATA_FILE):
        return {"users": [], "courses": []}
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            content = f.read().strip()
            if not content:
                return {"users": [], "courses": []}
            return json.loads(content)
    except Exception:
        return {"users": [], "courses": []}


def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def current_user():
    return session.get("user")


def require_login():
    if not session.get("user"):
        return jsonify({"error": "Not logged in"}), 401
    return None


def require_teacher():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401
    if user.get("role") != "teacher":
        return jsonify({"error": "Teacher access required"}), 403
    return None


def translate_to_hindi(text):
    try:
        if not text or not text.strip():
            return ""
        return GoogleTranslator(source="en", target="hi").translate(text) or ""
    except Exception:
        return ""


def translate_to_english(text):
    try:
        if not text or not text.strip():
            return ""
        return GoogleTranslator(source="hi", target="en").translate(text) or ""
    except Exception:
        return ""


def preprocess_image(img):
    img = img.convert("L")
    img = ImageEnhance.Contrast(img).enhance(2.5)
    img = ImageEnhance.Sharpness(img).enhance(2.0)
    img = img.filter(ImageFilter.MedianFilter(size=3))
    return img


def split_image(filepath):
    img = Image.open(filepath)
    width, height = img.size
    separator_x = find_separator(filepath)
    left = img.crop((0, 0, separator_x - 10, height))
    right = img.crop((separator_x + 10, 0, width, height))
    os.makedirs("uploads", exist_ok=True)
    left_path  = "uploads/left_en.png"
    right_path = "uploads/right_hi.png"
    preprocess_image(left).save(left_path)
    preprocess_image(right).save(right_path)
    return left_path, right_path


def extract_text_from_image(filepath, lang):
    try:
        img = Image.open(filepath)
        custom_config = r'--oem 3 --psm 6'
        text = pytesseract.image_to_string(img, lang=lang, config=custom_config)
        return text
    except Exception as e:
        print(f"OCR Error ({lang}): {e}")
        return ""


def extract_text_from_pdf_images(filepath):
    pages = convert_from_path(filepath, dpi=300)
    en_text = ""
    hi_text = ""
    os.makedirs("uploads/pdf_pages", exist_ok=True)
    for i, page in enumerate(pages):
        page_path = f"uploads/pdf_pages/page_{i}.png"
        page.save(page_path, "PNG")
        left_path, right_path = split_image(page_path)
        en_text += "\n" + extract_text_from_image(left_path, "eng")
        hi_text += "\n" + extract_text_from_image(right_path, "hin")
    return en_text, hi_text


def extract_text_from_pdf(filepath):
    text = ""
    try:
        with open(filepath, "rb") as f:
            reader = PyPDF2.PdfReader(f)
            for page in reader.pages:
                text += page.extract_text() + "\n"
    except Exception:
        text = ""
    return text


def extract_text_from_docx(filepath):
    text = ""
    try:
        doc = docx.Document(filepath)
        for para in doc.paragraphs:
            text += para.text + "\n"
    except Exception:
        text = ""
    return text


def extract_text_from_file(filepath, filename):
    ext = filename.rsplit(".", 1)[-1].lower()
    if ext in ["jpg", "jpeg", "png", "webp"]:
        left_path, right_path = split_image(filepath)
        en_text = extract_text_from_image(left_path, lang="eng")
        hi_text = extract_text_from_image(right_path, lang="hin")
        return en_text, hi_text
    elif ext == "pdf":
        try:
            return extract_text_from_pdf_images(filepath)
        except Exception as e:
            print("PDF OCR fallback:", e)
        text = extract_text_from_pdf(filepath)
        return text, ""
    elif ext in ["doc", "docx"]:
        text = extract_text_from_docx(filepath)
        return text, ""
    return "", ""


def clean_english(text):
    text = text.replace("|", "I").replace("'", "'")
    text = re.sub(r"[~`@#$%^&*_+=<>{}\\]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    doc = nlp_en(text)
    return " ".join([t.text for t in doc if not t.is_space])


def clean_hindi(text):
    text = re.sub(r"[^\u0900-\u097F\s\?\-।,०-९]", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def parse_questions(text, language="en"):
    questions = []
    if language == "hi":
        text = re.sub(r"\(6\)",   "(A)", text)
        text = re.sub(r"\(8\)\.", "(B)", text)
        text = re.sub(r"\(8\)",   "(B)", text)
        text = re.sub(r"\(0\)\.", "(C)", text)
        text = re.sub(r"\(0\)",   "(C)", text)
        text = re.sub(r"_#",      "(A)", text)
        text = re.sub(r"\(©\)",   "(C)", text)
        text = re.sub(r"\(o\)",   "(C)", text)
    text = re.sub(r"\(([A-Da-d])\)", r"(\1) ", text)
    blocks = re.split(r"\n(?=\d{1,3}[\.\)]\s)", text.strip())
    for block in blocks:
        lines = [l.strip() for l in block.splitlines() if l.strip()]
        if len(lines) < 2:
            continue
        num_match = re.match(r"^(\d{1,3})[\.\)]\s*(.*)", lines[0])
        if not num_match:
            continue
        q_num = int(num_match.group(1))
        if q_num > 200:
            continue
        question_text = num_match.group(2).strip()
        options = {}
        for line in lines[1:]:
            opt = re.match(r"^\(?([A-Da-d])\)?\s*[\.\)]?\s*(.+)", line)
            if opt:
                label = opt.group(1).upper()
                val   = opt.group(2).strip()
                val = re.sub(r"\s*\([A-D]\)\s*$", "", val).strip()
                if language == "hi":
                    val = clean_hindi(val)
                else:
                    val = clean_english(val)
                if val and label in ["A","B","C","D"]:
                    options[label] = val
            else:
                if not options and line:
                    question_text += " " + line
        if language == "hi":
            question_text = clean_hindi(question_text)
        else:
            question_text = clean_english(question_text)
        question_text = re.sub(r"^\d+[\.\)]\s*", "", question_text).strip()
        if len(options) >= 2 and len(question_text) > 5:
            questions.append({
                "q_num":    q_num,
                "question": question_text,
                "options":  options
            })
    return questions


def pair_and_translate(en_list, hi_list):
    hi_dict = {q["q_num"]: q for q in hi_list}
    paired  = []
    for i, en in enumerate(en_list):
        hi = hi_dict.get(en["q_num"])
        en_trans = {
            "question":    en["question"],
            "options":     en["options"],
            "explanation": "",
            "translated":  False
        }
        if hi and hi.get("question"):
            hi_trans = {
                "question":    hi["question"],
                "options":     hi["options"],
                "explanation": "",
                "translated":  False
            }
        else:
            hi_question = translate_to_hindi(en["question"])
            hi_options  = {lbl: translate_to_hindi(opt) for lbl, opt in en["options"].items()}
            hi_trans = {
                "question":    hi_question,
                "options":     hi_options,
                "explanation": "",
                "translated":  True
            }
        paired.append({
            "id":              i + 1,
            "question_number": en["q_num"],
            "question_type":   "multiple_choice",
            "translations":    {"en": en_trans, "hi": hi_trans},
            "correct_answer":  "",
            "difficulty":      "medium",
            "marks":           1,
            "subject":         "",
            "tags":            [],
            "hint":            "",
            "uploaded_by":     current_user(),
            "uploaded_at":     now()
        })
    return paired

def autotag_difficulty(question_text, options):
    """
    Rule-based difficulty tagger (no external AI needed).
    Easy:   short question, simple vocabulary
    Hard:   long question or advanced keywords
    Medium: everything else
    """
    text = question_text.lower()
    hard_keywords = ["calculate", "derive", "prove", "analyse", "evaluate",
                     "distinguish", "differentiate", "complex", "formula",
                     "theorem", "equation", "mechanism", "synthesis"]
    easy_keywords = ["what is", "who is", "which", "name the", "define",
                     "capital of", "full form", "abbreviation"]
    word_count = len(text.split())
    if any(k in text for k in hard_keywords) or word_count > 25:
        return "hard"
    if any(k in text for k in easy_keywords) or word_count < 10:
        return "easy"
    return "medium"

@app.route("/")
def index():
    return render_template("frontend.html")

@app.route("/api/register", methods=["POST"])
def register():
    body     = request.get_json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()
    name     = body.get("name", "").strip()
    role     = body.get("role", "teacher").strip()   # Feature 10: student/teacher

    if not username or not password or not name:
        return jsonify({"error": "All fields required"}), 400

    if role not in ("teacher", "student"):
        role = "teacher"

    data = load_data()
    if any(u["username"] == username for u in data["users"]):
        return jsonify({"error": "Username already exists"}), 400

    user = {
        "id":         len(data["users"]) + 1,
        "username":   username,
        "password":   generate_password_hash(password, method='pbkdf2:sha256'),
        "name":       name,
        "role":       role,
        "created_at": now()
    }
    data["users"].append(user)
    save_data(data)
    return jsonify({"success": True})

@app.route("/api/login", methods=["POST"])
def login():
    body     = request.get_json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()

    data = load_data()
    user = next((u for u in data["users"] if u["username"] == username), None)

    if not user or not check_password_hash(user["password"], password):
        return jsonify({"error": "Invalid username or password"}), 401

    session["user"] = {
        "id":       user["id"],
        "username": user["username"],
        "name":     user["name"],
        "role":     user.get("role", "teacher")
    }
    return jsonify({"success": True, "user": session["user"]})

@app.route("/api/logout", methods=["POST"])
def logout():
    session.clear()
    return jsonify({"success": True})

@app.route("/api/me", methods=["GET"])
def me():
    user = session.get("user")
    if not user:
        return jsonify({"error": "Not logged in"}), 401
    return jsonify(user)

@app.route("/api/users", methods=["GET"])
def get_users():
    data = load_data()
    return jsonify([
        {"id": u["id"], "username": u["username"], "name": u["name"], "role": u.get("role","teacher")}
        for u in data["users"]
    ])

@app.route("/api/courses", methods=["POST"])
def create_course():
    err = require_login()
    if err: return err
    body   = request.get_json()
    data   = load_data()
    user   = current_user()
    course = {
        "id":          len(data["courses"]) + 1,
        "name":        body.get("name", ""),
        "code":        body.get("code", ""),
        "subject":     body.get("subject", ""),
        "difficulty":  body.get("difficulty", "medium"),
        "marks":       body.get("marks", 1),
        "description": body.get("description", ""),
        "skills":      [],
        "lessons":     [],
        "question_sets": [],
        "created_by":  user["name"],
        "created_by_id": user["id"],
        "created_at":  now(),
        "updated_by":  user["name"],
        "updated_at":  now()
    }
    data["courses"].append(course)
    save_data(data)
    return jsonify({"success": True, "course": course})

@app.route("/api/courses", methods=["GET"])
def get_courses():
    data = load_data()
    return jsonify(data.get("courses", []))

@app.route("/api/courses/<int:course_id>", methods=["GET"])
def get_course(course_id):
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            return jsonify(c)
    return jsonify({"error": "Not found"}), 404

@app.route("/api/courses/<int:course_id>", methods=["PUT"])
def update_course(course_id):
    err = require_login()
    if err: return err
    body = request.get_json()
    data = load_data()
    user = current_user()
    for i, c in enumerate(data["courses"]):
        if c["id"] == course_id:
            c.update({
                "name":        body.get("name", c["name"]),
                "code":        body.get("code", c["code"]),
                "subject":     body.get("subject", c["subject"]),
                "difficulty":  body.get("difficulty", c["difficulty"]),
                "marks":       body.get("marks", c["marks"]),
                "description": body.get("description", c["description"]),
                "updated_by":  user["name"],
                "updated_at":  now()
            })
            data["courses"][i] = c
            save_data(data)
            return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route("/api/courses/<int:course_id>/skills", methods=["POST"])
def add_skill(course_id):
    err = require_login()
    if err: return err
    data  = load_data()
    user  = current_user()
    body  = request.get_json()
    for c in data["courses"]:
        if c["id"] == course_id:
            skill = {
                "id":          len(c["skills"]) + 1,
                "name":        body.get("name", ""),
                "description": body.get("description", ""),
                "added_by":    user["name"],
                "added_at":    now()
            }
            c["skills"].append(skill)
            c["updated_by"] = user["name"]
            c["updated_at"] = now()
            save_data(data)
            return jsonify({"success": True, "skill": skill})
    return jsonify({"error": "Course not found"}), 404

@app.route("/api/courses/<int:course_id>/skills", methods=["GET"])
def get_skills(course_id):
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            return jsonify(c.get("skills", []))
    return jsonify([])

@app.route("/api/courses/<int:course_id>/lessons", methods=["POST"])
def add_lesson(course_id):
    err = require_login()
    if err: return err
    data = load_data()
    user = current_user()
    body = request.get_json()
    for c in data["courses"]:
        if c["id"] == course_id:
            lesson = {
                "id":       len(c["lessons"]) + 1,
                "sl_no":    body.get("sl_no", len(c["lessons"]) + 1),
                "name":     body.get("name", ""),
                "summary":  body.get("summary", ""),
                "added_by": user["name"],
                "added_at": now()
            }
            c["lessons"].append(lesson)
            c["updated_by"] = user["name"]
            c["updated_at"] = now()
            save_data(data)
            return jsonify({"success": True, "lesson": lesson})
    return jsonify({"error": "Course not found"}), 404

@app.route("/api/courses/<int:course_id>/lessons", methods=["GET"])
def get_lessons(course_id):
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            return jsonify(c.get("lessons", []))
    return jsonify([])

@app.route("/api/courses/<int:course_id>/question_sets", methods=["POST"])
def save_question_set(course_id):
    err = require_login()
    if err: return err
    data = load_data()
    user = current_user()
    body = request.get_json()
    qset = body.get("question_set", {})

    for c in data["courses"]:
        if c["id"] == course_id:
            qset["id"]          = len(c["question_sets"]) + 1
            qset["created_by"]  = user["name"]
            qset["created_at"]  = now()
            qset["updated_by"]  = user["name"]
            qset["updated_at"]  = now()
            qset["edit_history"] = [{"action": "created", "by": user["name"], "at": now()}]
            c["question_sets"].append(qset)
            c["updated_by"] = user["name"]
            c["updated_at"] = now()
            save_data(data)
            return jsonify({"success": True, "id": qset["id"]})
    return jsonify({"error": "Course not found"}), 404

@app.route("/api/courses/<int:course_id>/question_sets", methods=["GET"])
def get_question_sets(course_id):
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            return jsonify(c.get("question_sets", []))
    return jsonify([])

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>", methods=["GET"])
def get_question_set(course_id, set_id):
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            for qs in c.get("question_sets", []):
                if qs["id"] == set_id:
                    return jsonify(qs)
    return jsonify({"error": "Not found"}), 404

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>", methods=["PUT"])
def update_question_set(course_id, set_id):
    err = require_login()
    if err: return err
    data = load_data()
    user = current_user()
    body = request.get_json()

    for c in data["courses"]:
        if c["id"] == course_id:
            for i, qs in enumerate(c["question_sets"]):
                if qs["id"] == set_id:
                    body["id"]         = set_id
                    body["created_by"] = qs.get("created_by", user["name"])
                    body["created_at"] = qs.get("created_at", now())
                    body["updated_by"] = user["name"]
                    body["updated_at"] = now()
                    history = qs.get("edit_history", [])
                    history.append({"action": "edited", "by": user["name"], "at": now()})
                    body["edit_history"] = history
                    c["question_sets"][i] = body
                    c["updated_by"] = user["name"]
                    c["updated_at"] = now()
                    save_data(data)
                    return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>", methods=["DELETE"])
def delete_question_set(course_id, set_id):
    err = require_login()
    if err: return err
    data = load_data()
    for c in data["courses"]:
        if c["id"] == course_id:
            c["question_sets"] = [qs for qs in c["question_sets"] if qs["id"] != set_id]
            save_data(data)
            return jsonify({"success": True})
    return jsonify({"error": "Not found"}), 404

@app.route("/upload", methods=["POST"])
def upload():
    try:
        files = request.files.getlist("files")
        if not files:
            return jsonify({"error": "No files uploaded"}), 400

        os.makedirs("uploads", exist_ok=True)
        all_en_questions = []
        all_hi_questions = []

        for file in files:
            if not file.filename:
                continue
            filepath = os.path.join("uploads", file.filename)
            file.save(filepath)
            en_text, hi_text = extract_text_from_file(filepath, file.filename)
            en_qs = parse_questions(en_text, language="en")
            hi_qs = parse_questions(hi_text, language="hi") if hi_text else []
            offset = len(all_en_questions)
            for q in en_qs:
                q["q_num"] += offset
            for q in hi_qs:
                q["q_num"] += offset
            all_en_questions.extend(en_qs)
            all_hi_questions.extend(hi_qs)

        if not all_en_questions:
            return jsonify({"error": "No questions found in uploaded files."}), 400

        paired = pair_and_translate(all_en_questions, all_hi_questions)

        frontend_questions = []
        for q in paired:
            en = q["translations"]["en"]
            hi = q["translations"]["hi"]
            # Feature 6: auto-tag difficulty
            diff = autotag_difficulty(en["question"], list(en["options"].values()))
            frontend_questions.append({
                "id":             q["id"],
                "question":       en["question"],
                "options":        list(en["options"].values()),
                "hindi_question": hi["question"],
                "hindi_options":  list(hi["options"].values()),
                "hi_translated":  hi["translated"],
                "en_translated":  en["translated"],
                "answer":         -1,
                "difficulty":     diff,   # Feature 6
                "hint":           "",     # Feature 7
                "explanation":    ""      # Feature 8
            })

        return jsonify({"success": True, "questions": frontend_questions})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>/add_questions", methods=["POST"])
def add_more_questions(course_id, set_id):
    err = require_login()
    if err: return err

    try:
        files = request.files.getlist("files")
        if not files:
            return jsonify({"error": "No files uploaded"}), 400

        os.makedirs("uploads", exist_ok=True)
        all_en_questions = []
        all_hi_questions = []

        for file in files:
            if not file.filename:
                continue
            filepath = os.path.join("uploads", file.filename)
            file.save(filepath)
            en_text, hi_text = extract_text_from_file(filepath, file.filename)
            all_en_questions.extend(parse_questions(en_text, language="en"))
            if hi_text:
                all_hi_questions.extend(parse_questions(hi_text, language="hi"))

        if not all_en_questions:
            return jsonify({"error": "No questions found."}), 400

        new_questions = pair_and_translate(all_en_questions, all_hi_questions)

        frontend_questions = []
        for q in new_questions:
            en = q["translations"]["en"]
            hi = q["translations"]["hi"]
            diff = autotag_difficulty(en["question"], list(en["options"].values()))
            frontend_questions.append({
                "id":             q["id"],
                "question":       en["question"],
                "options":        list(en["options"].values()),
                "hindi_question": hi["question"],
                "hindi_options":  list(hi["options"].values()),
                "hi_translated":  hi["translated"],
                "en_translated":  en["translated"],
                "answer":         -1,
                "difficulty":     diff,
                "hint":           "",
                "explanation":    ""
            })

        return jsonify({"success": True, "questions": frontend_questions})

    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/translate", methods=["POST"])
def translate():
    try:
        body      = request.get_json()
        text      = body.get("text", "")
        direction = body.get("direction", "en_to_hi")
        result    = translate_to_hindi(text) if direction == "en_to_hi" else translate_to_english(text)
        return jsonify({"success": True, "translated": result})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit", methods=["POST"])
def submit():
    body         = request.get_json()
    user_answers = body.get("answers", {})
    questions    = body.get("questions", [])
    score        = 0
    results      = []

    for i, q in enumerate(questions):
        user_ans      = user_answers.get(str(i), -1)
        correct_label = q.get("answer", "")
        correct_index = ["A","B","C","D"].index(correct_label) if correct_label in ["A","B","C","D"] else -1
        is_correct    = correct_index >= 0 and user_ans == correct_index
        if is_correct:
            score += 1
        results.append({
            "question":       q.get("question",""),
            "options":        q.get("options",[]),
            "user_answer":    user_ans,
            "correct_answer": correct_index,
            "is_correct":     is_correct,
            "explanation":    q.get("explanation", "")   # Feature 8
        })

    total = len(questions)
    return jsonify({
        "score":      score,
        "total":      total,
        "percentage": round((score / total) * 100) if total else 0,
        "results":    results
    })

@app.route("/api/attempts", methods=["POST"])
def save_attempt():
    """Save a quiz attempt after submission."""
    err = require_login()
    if err: return err

    body = request.get_json()
    data = load_data()

    if "attempts" not in data:
        data["attempts"] = []

    attempt = {
        "id":          len(data["attempts"]) + 1,
        "user_id":     current_user()["id"],
        "user_name":   current_user()["name"],
        "course_id":   body.get("course_id"),
        "course_name": body.get("course_name", ""),
        "set_id":      body.get("set_id"),
        "set_name":    body.get("set_name", ""),
        "score":       body.get("score", 0),
        "total":       body.get("total", 0),
        "percentage":  body.get("percentage", 0),
        "time_taken":  body.get("time_taken", 0),   # seconds
        "submitted_at": now()
    }
    data["attempts"].append(attempt)
    save_data(data)
    return jsonify({"success": True, "attempt": attempt})

@app.route("/api/attempts", methods=["GET"])
def get_attempts():
    err = require_login()
    if err: return err

    data     = load_data()
    user     = current_user()
    attempts = data.get("attempts", [])

    # Students see only their own; teachers see all
    if user.get("role") == "student":
        attempts = [a for a in attempts if a["user_id"] == user["id"]]

    # Optional filters
    course_id = request.args.get("course_id", type=int)
    set_id    = request.args.get("set_id", type=int)
    if course_id:
        attempts = [a for a in attempts if a.get("course_id") == course_id]
    if set_id:
        attempts = [a for a in attempts if a.get("set_id") == set_id]

    return jsonify(sorted(attempts, key=lambda x: x["submitted_at"], reverse=True))

@app.route("/api/search", methods=["GET"])
def search_questions():
    keyword = request.args.get("q", "").strip().lower()
    course_id = request.args.get("course_id", type=int)

    if not keyword:
        return jsonify({"results": []})

    data    = load_data()
    results = []

    for c in data["courses"]:
        if course_id and c["id"] != course_id:
            continue
        for qs in c.get("question_sets", []):
            for q in qs.get("questions", []):
                en_q = q.get("translations", {}).get("en", {}).get("question", "").lower()
                hi_q = q.get("translations", {}).get("hi", {}).get("question", "").lower()
                en_opts = " ".join(q.get("translations", {}).get("en", {}).get("options", {}).values()).lower()
                if keyword in en_q or keyword in hi_q or keyword in en_opts:
                    results.append({
                        "course_id":   c["id"],
                        "course_name": c["name"],
                        "set_id":      qs["id"],
                        "set_name":    qs.get("name", ""),
                        "question":    q.get("translations", {}).get("en", {}).get("question", ""),
                        "options":     list(q.get("translations", {}).get("en", {}).get("options", {}).values()),
                        "answer":      q.get("correct_answer", ""),
                        "difficulty":  q.get("difficulty", ""),
                        "hint":        q.get("hint", ""),
                    })

    return jsonify({"results": results, "count": len(results)})

@app.route("/api/autotag", methods=["POST"])
def autotag():
    """Auto-tag difficulty for a list of questions."""
    err = require_login()
    if err: return err

    body      = request.get_json()
    questions = body.get("questions", [])
    tagged    = []

    for q in questions:
        text    = q.get("question", "")
        options = q.get("options", [])
        diff    = autotag_difficulty(text, options)
        tagged.append({"id": q.get("id"), "difficulty": diff})

    return jsonify({"success": True, "tagged": tagged})

@app.route("/api/analytics", methods=["GET"])
def analytics():
    err = require_login()
    if err: return err

    data     = load_data()
    user     = current_user()
    attempts = data.get("attempts", [])

    if user.get("role") == "student":
        attempts = [a for a in attempts if a["user_id"] == user["id"]]

    course_id = request.args.get("course_id", type=int)
    if course_id:
        attempts = [a for a in attempts if a.get("course_id") == course_id]

    if not attempts:
        return jsonify({
            "total_attempts": 0,
            "avg_score": 0,
            "pass_rate": 0,
            "by_set": [],
            "by_student": [],
            "recent": []
        })

    total    = len(attempts)
    avg_pct  = round(sum(a["percentage"] for a in attempts) / total, 1)
    pass_cnt = sum(1 for a in attempts if a["percentage"] >= 60)
    pass_rate = round((pass_cnt / total) * 100, 1)

    set_map = {}
    for a in attempts:
        key = (a.get("set_id"), a.get("set_name",""))
        if key not in set_map:
            set_map[key] = {"attempts": 0, "total_pct": 0}
        set_map[key]["attempts"] += 1
        set_map[key]["total_pct"] += a["percentage"]

    by_set = [
        {
            "set_id":   k[0],
            "set_name": k[1],
            "attempts": v["attempts"],
            "avg_score": round(v["total_pct"] / v["attempts"], 1)
        }
        for k, v in set_map.items()
    ]

    by_student = []
    if user.get("role") == "teacher":
        stu_map = {}
        for a in attempts:
            uid = a["user_id"]
            if uid not in stu_map:
                stu_map[uid] = {"name": a["user_name"], "attempts": 0, "total_pct": 0}
            stu_map[uid]["attempts"] += 1
            stu_map[uid]["total_pct"] += a["percentage"]
        by_student = [
            {
                "user_id":  uid,
                "name":     v["name"],
                "attempts": v["attempts"],
                "avg_score": round(v["total_pct"] / v["attempts"], 1)
            }
            for uid, v in stu_map.items()
        ]

    recent = sorted(attempts, key=lambda x: x["submitted_at"], reverse=True)[:10]

    return jsonify({
        "total_attempts": total,
        "avg_score":      avg_pct,
        "pass_rate":      pass_rate,
        "by_set":         by_set,
        "by_student":     by_student,
        "recent":         recent
    })

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>/export/pdf")
def export_pdf(course_id, set_id):
    if not REPORTLAB_OK:
        return jsonify({"error": "reportlab not installed. Run: pip install reportlab"}), 500

    data = load_data()
    course = next((c for c in data["courses"] if c["id"] == course_id), None)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    qset = next((s for s in course.get("question_sets", []) if s["id"] == set_id), None)
    if not qset:
        return jsonify({"error": "Question set not found"}), 404

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)
    styles  = getSampleStyleSheet()
    story   = []
    LABELS  = ["A", "B", "C", "D"]

    title_style = ParagraphStyle("title", parent=styles["Heading1"],
                                 fontSize=16, textColor=colors.HexColor("#7c3aed"),
                                 spaceAfter=4)
    meta_style  = ParagraphStyle("meta", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.grey, spaceAfter=12)
    q_style     = ParagraphStyle("q", parent=styles["Normal"],
                                 fontSize=11, leading=15, spaceAfter=4)
    opt_style   = ParagraphStyle("opt", parent=styles["Normal"],
                                 fontSize=10, leading=13, leftIndent=15, spaceAfter=2)
    ans_style   = ParagraphStyle("ans", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#16a34a"),
                                 spaceAfter=6)
    hint_style  = ParagraphStyle("hint", parent=styles["Normal"],
                                 fontSize=9, textColor=colors.HexColor("#7c3aed"),
                                 spaceAfter=4)

    story.append(Paragraph(f"Question Paper: {qset.get('name','')}", title_style))
    story.append(Paragraph(
        f"Course: {course['name']}  |  Subject: {course.get('subject','')}  |  "
        f"Total Questions: {len(qset.get('questions',[]))}  |  "
        f"Created by: {qset.get('created_by','')}",
        meta_style))
    story.append(Spacer(1, 0.3*cm))

    for i, q in enumerate(qset.get("questions", []), 1):
        en = q.get("translations", {}).get("en", {})
        q_text   = en.get("question", "")
        options  = en.get("options", {})
        correct  = q.get("correct_answer", "")
        hint     = q.get("hint", "")
        expl     = en.get("explanation", "")
        diff     = q.get("difficulty", "")

        diff_color = {"easy": "#16a34a", "hard": "#dc2626"}.get(diff, "#7c3aed")
        diff_badge = f'<font color="{diff_color}">[{diff.upper()}]</font>' if diff else ""

        story.append(Paragraph(f"<b>Q{i}.</b> {q_text} {diff_badge}", q_style))
        for lbl, val in options.items():
            marker = "✓ " if lbl == correct else ""
            story.append(Paragraph(f"({lbl}) {marker}{val}", opt_style))

        if correct:
            story.append(Paragraph(f"Answer: {correct}", ans_style))
        if hint:
            story.append(Paragraph(f"💡 Hint: {hint}", hint_style))
        if expl:
            story.append(Paragraph(f"📖 Explanation: {expl}", hint_style))
        story.append(Spacer(1, 0.3*cm))

    doc.build(story)
    buf.seek(0)
    filename = f"{qset.get('name','questions').replace(' ','_')}.pdf"
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True, download_name=filename)

@app.route("/api/courses/<int:course_id>/question_sets/<int:set_id>/export/excel")
def export_excel(course_id, set_id):
    if not OPENPYXL_OK:
        return jsonify({"error": "openpyxl not installed. Run: pip install openpyxl"}), 500

    data = load_data()
    course = next((c for c in data["courses"] if c["id"] == course_id), None)
    if not course:
        return jsonify({"error": "Course not found"}), 404

    qset = next((s for s in course.get("question_sets", []) if s["id"] == set_id), None)
    if not qset:
        return jsonify({"error": "Question set not found"}), 404

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Questions"

    purple = PatternFill("solid", fgColor="7C3AED")
    white_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align   = Alignment(vertical="top", wrap_text=True)

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = ["#", "Question (EN)", "Question (HI)",
               "Option A", "Option B", "Option C", "Option D",
               "Correct Answer", "Difficulty", "Hint", "Explanation", "Marks"]
    widths  = [5, 40, 40, 20, 20, 20, 20, 14, 12, 25, 30, 8]

    for col, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill   = purple
        cell.font   = white_font
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 30

    LABELS = ["A", "B", "C", "D"]
    alt_fill = PatternFill("solid", fgColor="F5F3FF")

    for i, q in enumerate(qset.get("questions", []), 1):
        en      = q.get("translations", {}).get("en", {})
        hi      = q.get("translations", {}).get("hi", {})
        opts_en = en.get("options", {})
        row_fill = alt_fill if i % 2 == 0 else None

        row = [
            i,
            en.get("question", ""),
            hi.get("question", ""),
            opts_en.get("A", ""),
            opts_en.get("B", ""),
            opts_en.get("C", ""),
            opts_en.get("D", ""),
            q.get("correct_answer", ""),
            q.get("difficulty", ""),
            q.get("hint", ""),
            en.get("explanation", ""),
            q.get("marks", 1)
        ]
        for col, val in enumerate(row, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.alignment = cell_align
            cell.border    = border
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[i + 1].height = 45

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{qset.get('name','questions').replace(' ','_')}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=filename)

@app.route("/debug_ocr", methods=["POST"])
def debug_ocr():
    try:
        file = request.files.get("file")
        if not file:
            return jsonify({"error": "No file"}), 400
        os.makedirs("uploads", exist_ok=True)
        filepath = os.path.join("uploads", file.filename)
        file.save(filepath)
        left_path, right_path = split_image(filepath)
        en_text = extract_text_from_image(left_path, lang="eng")
        hi_text = extract_text_from_image(right_path, lang="hin")
        en_questions = parse_questions(en_text, language="en")
        hi_questions = parse_questions(hi_text, language="hi")
        return jsonify({
            "en_raw": en_text,
            "hi_raw": hi_text,
            "en_questions_found": len(en_questions),
            "hi_questions_found": len(hi_questions),
            "en_questions": en_questions,
            "hi_questions": hi_questions
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    import os, threading, webbrowser
    port = int(os.environ.get("PORT", 3000))
    threading.Timer(1.5, lambda: webbrowser.open(f"http://localhost:{port}")).start()
    app.run(host="0.0.0.0", port=port, debug=True)